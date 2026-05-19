"""
블로그 아티클의 예제 3개를 파싱해서 독자가 바로 연습할 수 있는 Excel 파일을 생성합니다.
아티클 .txt 파일과 같은 디렉터리에 _practice.xlsx 파일을 저장합니다.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── 예제 섹션 헤더로 쓰이는 레이블 (refine_drafts_ai.py STEP_LABELS / SCENARIO_LABELS 와 동기화)
_STEP_LABEL_KO = {"단계별 방법", "적용 순서", "실행 단계", "진행 방식", "확인 방법"}
_SCENARIO_LABEL_PREFIX = re.compile(
    r"^(?:예제|체크 항목|비교 장면|처음 효과 장면|Check item|Comparison scene|First-try scene)"
    r"\s+(\d+)\.\s+(.+?)(?:\s*\[.*?\])?\s*$"
)

# 각 예제 내부 파트를 인식하는 패턴
_SITUATION_START = re.compile(r"^상황\s*설명\s*:")
_WHY_START = re.compile(r"^(?:왜\s+이\s+순서가\s+중요한가|왜\s+필요한가)\s*:")
_CAUTION_START = re.compile(r"^주의사항\s*:")
_ENGLISH_LINE = re.compile(
    r"^(Situation:|Why this scene matters:|Why it matters:|Boundary caution:|Caution:|"
    r"Application sequence:|Execution steps:|How to proceed:|Step-by-step approach:|How to verify:)"
)
_STEP_BULLET = re.compile(r"^(?:-\s+|\d+[.)]\s+)(.+)$")


@dataclass
class ScenarioData:
    index: int
    title: str
    situation: str = ""
    why: str = ""
    steps: list[str] = field(default_factory=list)
    step_label: str = "적용 순서"
    caution: str = ""


def _is_english_paragraph(line: str) -> bool:
    """영문 단락 시작 여부 (한국어 연습 시트에서 제외)"""
    if _ENGLISH_LINE.match(line):
        return True
    # 영문 bullet (-으로 시작하지 않는 라틴 문자 비율이 높은 경우)
    latin = sum(1 for c in line if c.isascii() and c.isalpha())
    total = sum(1 for c in line if c.isalpha())
    return total > 0 and latin / total > 0.7


def _clean(text: str) -> str:
    return text.strip()


def parse_scenarios(article_text: str) -> list[ScenarioData]:
    """아티클 텍스트에서 예제 3개를 파싱해 ScenarioData 리스트로 반환합니다."""
    scenarios: list[ScenarioData] = []
    current: ScenarioData | None = None

    # 파싱 상태
    state = "idle"  # idle | situation | why | steps | caution
    buffer: list[str] = []
    skip_english = False

    def _flush_buffer(sc: ScenarioData, st: str, buf: list[str]):
        text = " ".join(buf).strip()
        if st == "situation":
            sc.situation = text
        elif st == "why":
            sc.why = text
        elif st == "caution":
            sc.caution = text

    for raw_line in article_text.splitlines():
        line = raw_line.strip()

        # 새로운 예제 헤더 감지
        m = _SCENARIO_LABEL_PREFIX.match(line)
        if m:
            if current and state != "idle":
                if state == "steps":
                    pass  # steps already appended inline
                else:
                    _flush_buffer(current, state, buffer)
            if current:
                scenarios.append(current)
            current = ScenarioData(index=int(m.group(1)), title=_clean(m.group(2)))
            state = "idle"
            buffer = []
            skip_english = False
            continue

        if current is None:
            continue

        # 영문 단락 스킵 (한국어 파트만 추출)
        if _is_english_paragraph(line):
            skip_english = True
            continue
        if skip_english:
            # 빈 줄이 나오면 영문 단락 끝
            if not line:
                skip_english = False
            continue

        # 상황 설명 시작
        if _SITUATION_START.match(line):
            _flush_buffer(current, state, buffer)
            state = "situation"
            buffer = [re.sub(r"^상황\s*설명\s*:\s*", "", line)]
            continue

        # 왜 이 순서 시작
        if _WHY_START.match(line):
            _flush_buffer(current, state, buffer)
            state = "why"
            buffer = [re.sub(r"^(?:왜\s+이\s+순서가\s+중요한가|왜\s+필요한가)\s*:\s*", "", line)]
            continue

        # 스텝 레이블 시작
        if line in _STEP_LABEL_KO and line.endswith(":") is False:
            # "적용 순서:" 형태 또는 그냥 "적용 순서"
            pass
        step_header = re.match(r"^(" + "|".join(re.escape(s) for s in _STEP_LABEL_KO) + r")\s*:?\s*$", line)
        if step_header:
            _flush_buffer(current, state, buffer)
            current.step_label = step_header.group(1)
            state = "steps"
            buffer = []
            continue

        # 주의사항 시작
        if _CAUTION_START.match(line):
            _flush_buffer(current, state, buffer)
            state = "caution"
            buffer = [re.sub(r"^주의사항\s*:\s*", "", line)]
            continue

        # 빈 줄
        if not line:
            if state in ("situation", "why", "caution"):
                _flush_buffer(current, state, buffer)
                state = "idle"
                buffer = []
            continue

        # 스텝 bullet 수집
        if state == "steps":
            bullet_match = _STEP_BULLET.match(line)
            if bullet_match:
                current.steps.append(bullet_match.group(1).strip())
            continue

        # 나머지 텍스트를 현재 버퍼에 추가
        if state in ("situation", "why", "caution"):
            buffer.append(line)

    # 마지막 예제 처리
    if current:
        if state != "steps":
            _flush_buffer(current, state, buffer)
        scenarios.append(current)

    return scenarios


# ── Excel 스타일 상수

_COLOR_HEADER_BG = "2F5496"   # 진한 파란색
_COLOR_SCENARIO_BG = "D6E4F0"  # 연한 파란색
_COLOR_STEP_BG = "EBF5FB"      # 아주 연한 파란색
_COLOR_PRACTICE_BG = "FDFEFE"  # 거의 흰색
_COLOR_CAUTION_BG = "FEF9E7"   # 연한 노란색
_COLOR_GUIDE_BG = "F0F3F4"     # 연한 회색

_FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
_FONT_SECTION = Font(name="맑은 고딕", size=11, bold=True, color="1A5276")
_FONT_BODY = Font(name="맑은 고딕", size=10)
_FONT_STEP_NUM = Font(name="맑은 고딕", size=10, bold=True, color="2F5496")
_FONT_PRACTICE_HINT = Font(name="맑은 고딕", size=9, italic=True, color="95A5A6")
_FONT_GUIDE_TITLE = Font(name="맑은 고딕", size=16, bold=True, color="2C3E50")
_FONT_GUIDE_BODY = Font(name="맑은 고딕", size=10, color="2C3E50")

_THIN = Side(style="thin", color="BDC3C7")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_cell(ws, row: int, col: int, value, font=None, fill=None,
              alignment=None, border=None, height: float | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if height is not None:
        ws.row_dimensions[row].height = height
    return cell


def _merge_row(ws, row: int, col_start: int, col_end: int,
               value, font=None, fill=None, alignment=None, height: float | None = None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    _set_cell(ws, row, col_start, value, font=font, fill=fill,
              alignment=alignment, height=height)
    for col in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=col).border = _BORDER_THIN


def _build_guide_sheet(wb: openpyxl.Workbook, title: str, scenarios: list[ScenarioData]):
    ws = wb.active
    ws.title = "사용 안내"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 4

    # 상단 타이틀
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 50
    ws.merge_cells("B2:C2")
    cell = ws.cell(row=2, column=2,
                   value=f"📘 {title}\n\n연습 워크시트 사용 안내")
    cell.font = _FONT_GUIDE_TITLE
    cell.fill = _fill("2C3E50")
    cell.alignment = _CENTER
    ws.cell(row=2, column=2).font = Font(name="맑은 고딕", size=15, bold=True, color="FFFFFF")

    # 안내 본문
    guide_rows = [
        ("", ""),
        ("이 파일 구성", f"총 {len(scenarios)}개의 시트 (예제 1 ~ 예제 {len(scenarios)})"),
        ("", "각 시트는 블로그 예제 하나에 대한 연습 공간입니다."),
        ("", ""),
        ("사용 방법", "1단계: 상황 설명을 읽고 어떤 맥락인지 파악합니다."),
        ("", "2단계: 적용 순서 각 단계를 직접 따라 해봅니다."),
        ("", "3단계: [내 답변 / 메모] 열에 자신의 방식으로 정리합니다."),
        ("", "4단계: 주의사항을 읽고 '나라면 어떻게 피할지' 적어봅니다."),
        ("", ""),
        ("팁", "빈 칸은 정답이 없습니다. 자신의 상황에 맞게 자유롭게 기록하세요."),
        ("", "작성 후 날짜를 적어두면 나중에 비교하기 좋습니다."),
    ]
    r = 4
    for label, content in guide_rows:
        ws.row_dimensions[r].height = 20 if content else 8
        if label:
            _set_cell(ws, r, 2, label,
                      font=_FONT_SECTION,
                      fill=_fill(_COLOR_GUIDE_BG),
                      alignment=_WRAP,
                      border=_BORDER_THIN)
            _set_cell(ws, r, 3, content,
                      font=_FONT_GUIDE_BODY,
                      fill=_fill(_COLOR_GUIDE_BG),
                      alignment=_WRAP,
                      border=_BORDER_THIN)
        elif content:
            ws.cell(row=r, column=2).border = _BORDER_THIN
            _set_cell(ws, r, 3, content,
                      font=_FONT_GUIDE_BODY,
                      fill=_fill(_COLOR_GUIDE_BG),
                      alignment=_WRAP,
                      border=_BORDER_THIN)
        r += 1

    # 예제 목록
    r += 1
    _merge_row(ws, r, 2, 3, "예제 목록",
               font=_FONT_SECTION,
               fill=_fill(_COLOR_SCENARIO_BG),
               alignment=_WRAP,
               height=22)
    r += 1
    for sc in scenarios:
        ws.row_dimensions[r].height = 22
        _set_cell(ws, r, 2, f"예제 {sc.index}",
                  font=_FONT_STEP_NUM,
                  fill=_fill(_COLOR_STEP_BG),
                  alignment=_CENTER,
                  border=_BORDER_THIN)
        _set_cell(ws, r, 3, sc.title,
                  font=_FONT_BODY,
                  fill=_fill(_COLOR_STEP_BG),
                  alignment=_WRAP,
                  border=_BORDER_THIN)
        r += 1


def _build_scenario_sheet(wb: openpyxl.Workbook, sc: ScenarioData, keyword: str):
    ws = wb.create_sheet(title=f"예제 {sc.index}")

    # 열 너비
    ws.column_dimensions["A"].width = 4    # 여백
    ws.column_dimensions["B"].width = 6    # 단계 번호
    ws.column_dimensions["C"].width = 42   # 내용 설명
    ws.column_dimensions["D"].width = 38   # 내 답변/메모
    ws.column_dimensions["E"].width = 4    # 여백

    r = 1
    ws.row_dimensions[r].height = 8
    r += 1

    # ── 예제 타이틀 행
    ws.row_dimensions[r].height = 44
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=2,
                   value=f"예제 {sc.index}.  {sc.title}")
    cell.font = _FONT_TITLE
    cell.fill = _fill(_COLOR_HEADER_BG)
    cell.alignment = _CENTER
    r += 1

    # ── 키워드 행
    ws.row_dimensions[r].height = 20
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=2,
                   value=f"키워드: {keyword}")
    cell.font = Font(name="맑은 고딕", size=9, color="7F8C8D")
    cell.fill = _fill("EAF2FF")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 2

    # ── 상황 설명
    if sc.situation:
        ws.row_dimensions[r].height = 20
        _merge_row(ws, r, 2, 4, "📌 상황 설명",
                   font=_FONT_SECTION,
                   fill=_fill(_COLOR_SCENARIO_BG),
                   alignment=_WRAP,
                   height=22)
        r += 1
        ws.row_dimensions[r].height = max(60, len(sc.situation) // 2)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell = ws.cell(row=r, column=2, value=sc.situation)
        cell.font = _FONT_BODY
        cell.fill = _fill("FAFAFA")
        cell.alignment = _WRAP
        cell.border = _BORDER_THIN
        r += 2

    # ── 단계별 연습 테이블
    ws.row_dimensions[r].height = 22
    _merge_row(ws, r, 2, 4,
               f"🔢 {sc.step_label}  —  직접 따라 해보세요",
               font=_FONT_SECTION,
               fill=_fill(_COLOR_SCENARIO_BG),
               alignment=_WRAP,
               height=22)
    r += 1

    # 테이블 헤더
    ws.row_dimensions[r].height = 22
    for col, (label, w) in enumerate(
        [("단계", 6), ("단계 설명 (원문)", 42), ("내 답변 / 메모 (직접 작성)", 38)],
        start=2,
    ):
        cell = ws.cell(row=r, column=col, value=label)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        cell.fill = _fill("5D6D7E")
        cell.alignment = _CENTER
        cell.border = _BORDER_THIN
    r += 1

    # 단계 rows
    for i, step_text in enumerate(sc.steps, start=1):
        step_height = max(36, len(step_text) // 1.5)
        ws.row_dimensions[r].height = step_height

        _set_cell(ws, r, 2, i,
                  font=_FONT_STEP_NUM,
                  fill=_fill(_COLOR_STEP_BG),
                  alignment=_CENTER,
                  border=_BORDER_THIN)
        _set_cell(ws, r, 3, step_text,
                  font=_FONT_BODY,
                  fill=_fill(_COLOR_STEP_BG),
                  alignment=_WRAP,
                  border=_BORDER_THIN)
        practice_cell = ws.cell(row=r, column=4)
        practice_cell.font = _FONT_PRACTICE_HINT
        practice_cell.fill = _fill(_COLOR_PRACTICE_BG)
        practice_cell.alignment = _WRAP
        practice_cell.border = _BORDER_THIN
        r += 1

    r += 1

    # ── 왜 이 순서인가
    if sc.why:
        _merge_row(ws, r, 2, 4, "💡 왜 이 순서가 중요한가",
                   font=_FONT_SECTION,
                   fill=_fill(_COLOR_SCENARIO_BG),
                   alignment=_WRAP,
                   height=22)
        r += 1
        height = max(55, len(sc.why) // 2)
        ws.row_dimensions[r].height = height
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell = ws.cell(row=r, column=2, value=sc.why)
        cell.font = _FONT_BODY
        cell.fill = _fill("FAFAFA")
        cell.alignment = _WRAP
        cell.border = _BORDER_THIN
        r += 2

    # ── 주의사항
    if sc.caution:
        _merge_row(ws, r, 2, 4, "⚠️ 주의사항",
                   font=Font(name="맑은 고딕", size=11, bold=True, color="C0392B"),
                   fill=_fill(_COLOR_CAUTION_BG),
                   alignment=_WRAP,
                   height=22)
        r += 1
        height = max(60, len(sc.caution) // 2)
        ws.row_dimensions[r].height = height
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell = ws.cell(row=r, column=2, value=sc.caution)
        cell.font = Font(name="맑은 고딕", size=10, color="922B21")
        cell.fill = _fill(_COLOR_CAUTION_BG)
        cell.alignment = _WRAP
        cell.border = _BORDER_THIN
        r += 2

    # ── 자유 메모 공간
    _merge_row(ws, r, 2, 4, "✏️ 자유 메모  (직접 써보세요)",
               font=_FONT_SECTION,
               fill=_fill(_COLOR_GUIDE_BG),
               alignment=_WRAP,
               height=22)
    r += 1
    ws.row_dimensions[r].height = 90
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    memo_cell = ws.cell(row=r, column=2)
    memo_cell.fill = _fill("FFFFFF")
    memo_cell.border = _BORDER_THIN
    memo_cell.alignment = _WRAP


def generate_practice_excel(article_path: str | Path) -> Path | None:
    """
    아티클 .txt 파일을 읽어 예제 3개 연습 Excel 파일을 생성합니다.
    생성된 .xlsx 경로를 반환하며, 예제 파싱 실패 시 None을 반환합니다.
    """
    article_path = Path(article_path)
    if not article_path.exists():
        print(f"[practice_excel] 파일 없음: {article_path}")
        return None

    text = article_path.read_text(encoding="utf-8")

    # 키워드 추출 (아티클 헤더에서)
    keyword = ""
    for line in text.splitlines():
        m = re.match(r"^핵심\s*키워드\s*:\s*(.+)$", line)
        if m:
            keyword = m.group(1).strip()
            break
    if not keyword:
        keyword = article_path.stem

    scenarios = parse_scenarios(text)
    if not scenarios:
        print(f"[practice_excel] 예제 파싱 실패: {article_path.name}")
        return None

    # 제목 추출
    title = keyword
    for line in text.splitlines():
        m = re.match(r"^제목\s*:\s*(.+)$", line)
        if m:
            title = m.group(1).strip()
            break

    wb = openpyxl.Workbook()
    _build_guide_sheet(wb, title, scenarios)
    for sc in scenarios:
        _build_scenario_sheet(wb, sc, keyword)

    out_path = article_path.with_name(article_path.stem + "_practice.xlsx")
    wb.save(out_path)
    print(f"[practice_excel] 저장 완료: {out_path.name}")
    return out_path


def generate_practice_excel_batch(article_paths: list[str | Path]) -> list[Path]:
    """여러 아티클에 대해 일괄 생성합니다."""
    results = []
    for p in article_paths:
        out = generate_practice_excel(p)
        if out:
            results.append(out)
    return results


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("사용법: python generate_practice_excel.py <article.txt> [article2.txt ...]")
        sys.exit(1)
    for arg in sys.argv[1:]:
        generate_practice_excel(arg)
